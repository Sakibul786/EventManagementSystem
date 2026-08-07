from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
)

from events.models import OfflineAttendance


def generate_attendance_excel(event):

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Attendance Report"

    # =====================================
    # Title
    # =====================================

    worksheet.merge_cells("A1:D1")

    title = worksheet["A1"]

    title.value = "EVENT ATTENDANCE REPORT"

    title.font = Font(
        size=16,
        bold=True,
    )

    title.alignment = Alignment(
        horizontal="center"
    )

    # =====================================
    # Event Information
    # =====================================

    worksheet["A3"] = "Event"
    worksheet["B3"] = event.name

    worksheet["A4"] = "Date"
    worksheet["B4"] = str(event.date)

    worksheet["A5"] = "Time"
    worksheet["B5"] = str(event.time)

    worksheet["A6"] = "Location"
    worksheet["B6"] = event.location

    if event.category:

        worksheet["A7"] = "Category"

        worksheet["B7"] = event.category.name

    # =====================================
    # Attendance Table
    # =====================================

    start_row = 9

    headers = [

        "No",

        "Participant",

        "Email",

        "Status",

    ]

    fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78",
    )

    for column, header in enumerate(
        headers,
        start=1,
    ):

        cell = worksheet.cell(
            row=start_row,
            column=column,
        )

        cell.value = header

        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )

        cell.fill = fill

        cell.alignment = Alignment(
            horizontal="center"
        )

    total_present = 0

    row = start_row + 1

    participants = event.participants.all()

    for index, participant in enumerate(
        participants,
        start=1,
    ):

        attendance = event.attendance_records.filter(
            participant=participant
        ).first()

        status = "Absent"

        if attendance and attendance.is_present:

            status = "Present"

            total_present += 1

        worksheet.cell(
            row=row,
            column=1,
        ).value = index

        worksheet.cell(
            row=row,
            column=2,
        ).value = participant.name

        worksheet.cell(
            row=row,
            column=3,
        ).value = participant.email

        worksheet.cell(
            row=row,
            column=4,
        ).value = status

        row += 1

    # =====================================
    # Summary
    # =====================================
        offline_attendance, created = OfflineAttendance.objects.get_or_create(
        event=event
    )

    # =====================================
    # Online Statistics
    # =====================================

    online_registered = participants.count()

    online_present = total_present

    online_absent = (
        online_registered -
        online_present
    )

    # =====================================
    # Offline Statistics
    # =====================================

    offline_present = offline_attendance.present

    offline_registered = event.offline_participants

    offline_absent = max(
        0,
        offline_registered - offline_present
    )


    # =====================================
    # Overall Statistics
    # =====================================

    total_registered = (
        online_registered +
        offline_registered
    )

    total_present = (
        online_present +
        offline_present
    )

    total_absent = (
        online_absent +
        offline_absent
    )

    attendance_percentage = (
        round(
            (total_present / total_registered) * 100,
            1,
        )
        if total_registered > 0
        else 0
    )

    row += 2

    summary_fill = PatternFill(
        fill_type="solid",
        start_color="D9EAD3",
        end_color="D9EAD3",
    )

    summary = [

        ("Online Registered", online_registered),

        ("Offline Registered", offline_registered),

        ("Total Registered", total_registered),

        ("Online Present", online_present),

        ("Offline Present", offline_present),

        ("Total Present", total_present),

        ("Online Absent", online_absent),

        ("Offline Absent", offline_absent),

        ("Total Absent", total_absent),

        ("Attendance Percentage", f"{attendance_percentage}%"),

    ]

    for title, value in summary:

        worksheet.cell(
            row=row,
            column=1,
        ).value = title

        worksheet.cell(
            row=row,
            column=2,
        ).value = value

        worksheet.cell(
            row=row,
            column=1,
        ).font = Font(
            bold=True,
        )

        worksheet.cell(
            row=row,
            column=1,
        ).fill = summary_fill

        row += 1

    # =====================================
    # Column Width
    # =====================================

    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 30
    worksheet.column_dimensions["C"].width = 35
    worksheet.column_dimensions["D"].width = 18

    # =====================================
    # Response
    # =====================================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response[
        "Content-Disposition"
    ] = (
        f'attachment; '
        f'filename="{event.name}_Attendance_Report.xlsx"'
    )

    workbook.save(response)

    return response